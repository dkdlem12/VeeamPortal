from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from datetime import date
from typing import Optional
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from app.collectors import get_collector

router = APIRouter(prefix="/history", tags=["history"])


@router.get("/jobs/{session_id}")
def get_session_detail(session_id: str):
    return get_collector().get_session_detail(session_id)


@router.get("/jobs")
def get_job_history(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    job_type: Optional[str] = None,
    status: Optional[str] = None,
    job_name: Optional[str] = None,
    server: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
):
    return get_collector().get_job_history(
        start_date, end_date, job_type, status, job_name, server, page, page_size
    )


@router.get("/export")
def export_job_history(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    job_type: Optional[str] = None,
    status: Optional[str] = None,
    job_name: Optional[str] = None,
    server: Optional[str] = None,
):
    result = get_collector().get_job_history(
        start_date, end_date, job_type, status, job_name, server, page=1, page_size=10000
    )
    jobs = result["items"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "백업 수행 이력"

    header_fill = PatternFill(start_color="1B6CA8", end_color="1B6CA8", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    BACKUP_MODE_KO = {
        "Incremental":   "증분",
        "SyntheticFull": "합성 풀",
        "ActiveFull":    "액티브 풀",
    }
    headers = [
        "Job 이름", "백업 유형", "백업 모드", "서버",
        "시작 시간", "종료 시간", "소요 시간(초)",
        "Processed(GB)", "Read(GB)", "Transferred(GB)",
        "상태",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, job in enumerate(jobs, 2):
        ws.cell(row=row_idx, column=1,  value=job.get("name", ""))
        ws.cell(row=row_idx, column=2,  value=job.get("type", ""))
        ws.cell(row=row_idx, column=3,  value=BACKUP_MODE_KO.get(job.get("backupMode", ""), job.get("backupMode", "")))
        ws.cell(row=row_idx, column=4,  value=job.get("server", ""))
        ws.cell(row=row_idx, column=5,  value=job.get("startTime", ""))
        ws.cell(row=row_idx, column=6,  value=job.get("endTime", "") or "")
        ws.cell(row=row_idx, column=7,  value=job.get("duration") or "")
        ws.cell(row=row_idx, column=8,  value=job.get("dataSize") or "")
        ws.cell(row=row_idx, column=9,  value=job.get("readSize") or "")
        ws.cell(row=row_idx, column=10, value=job.get("transferSize") or "")
        ws.cell(row=row_idx, column=11, value=job.get("status", ""))

    # Auto column width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"veeam_backup_history.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
